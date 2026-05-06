from __future__ import annotations

import argparse
from pathlib import Path
from openpyxl import load_workbook


def normalize_search_name(value: str) -> str:
    import unicodedata
    import re

    text = unicodedata.normalize("NFD", value or "")
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return re.sub(r"\s+", " ", text)


def sql_escape(value: str) -> str:
    return value.replace("'", "''")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_path")
    parser.add_argument("output_path")
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--default-series", default="")
    args = parser.parse_args()

    workbook = load_workbook(args.input_path, data_only=True)
    worksheet = workbook[args.sheet] if args.sheet else workbook.worksheets[0]

    rows = list(worksheet.iter_rows(values_only=True))
    header_index = None
    for index, row in enumerate(rows):
        values = [str(value).strip() if value is not None else "" for value in row]
        if "Nume și prenume" in values and "Grupa" in values and "Anul" in values:
            header_index = index
            break

    if header_index is None:
        raise ValueError("Nu am găsit antetele necesare în Excel.")

    header = [str(value).strip() if value is not None else "" for value in rows[header_index]]
    name_col = header.index("Nume și prenume")
    group_col = header.index("Grupa")
    year_col = header.index("Anul")

    inserts = []
    for row in rows[header_index + 1:]:
        full_name = str(row[name_col] or "").strip()
        group_code = str(row[group_col] or "").strip()
        study_year = str(row[year_col] or "").strip()

        if not full_name or not group_code or not study_year:
            continue

        search_name = normalize_search_name(full_name)
        series = args.default_series.strip()
        inserts.append(
            "('"
            + sql_escape(full_name)
            + "', '"
            + sql_escape(search_name)
            + "', null, '"
            + sql_escape(study_year)
            + "', '"
            + sql_escape(series)
            + "', '"
            + sql_escape(group_code)
            + "')"
        )

    output_lines = [
      "insert into public.students (full_name, search_name, email, study_year, series, group_code)",
      "values",
      ",\n".join(inserts),
      "on conflict do nothing;",
      "",
    ]

    Path(args.output_path).write_text("\n".join(output_lines), encoding="utf-8")


if __name__ == "__main__":
    main()
